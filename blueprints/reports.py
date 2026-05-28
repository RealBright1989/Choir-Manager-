from datetime import datetime, date
from flask import Blueprint, render_template, request, Response, current_app, send_from_directory
from utils import login_required, export_excel, export_pdf, get_hierarchical_report, generate_hierarchy_pdf, generate_hierarchy_excel
from models import db, Member, Payment, Attendance, Setting, Song, User
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _paths():
    root = current_app.root_path
    return root + "/static/logo_watermark.png"

bp = Blueprint("reports", __name__)


@bp.route("/reports", defaults={"section": "standard"}, methods=["GET"])
@bp.route("/reports/<section>")
@login_required
def reports(section="standard"):
    if section == "hierarchy":
        zone_f = request.args.get("zone", "")
        area_f = request.args.get("area", "")
        state_f = request.args.get("state", "")
        country_f = request.args.get("country", "")
        zones = [r[0] for r in Member.query.with_entities(Member.zone).filter(Member.zone.isnot(None), Member.zone != "").distinct().order_by(Member.zone).all()]
        areas = [r[0] for r in Member.query.with_entities(Member.area).filter(Member.area.isnot(None), Member.area != "").distinct().order_by(Member.area).all()]
        states = [r[0] for r in Member.query.with_entities(Member.state_of_origin).filter(Member.state_of_origin.isnot(None), Member.state_of_origin != "").distinct().order_by(Member.state_of_origin).all()]
        countries = [r[0] for r in Member.query.with_entities(Member.country).filter(Member.country.isnot(None), Member.country != "").distinct().order_by(Member.country).all()]
        flat, hierarchy, grand = get_hierarchical_report(zone_f, area_f, state_f, country_f)
        return render_template("reports.html", section="hierarchy", flat=flat, grand=grand,
                               zones=zones, areas=areas, states=states, countries=countries,
                               filters={"zone": zone_f, "area": area_f, "state": state_f, "country": country_f})

    member_financials = db.session.query(
        Member.first_name, Member.last_name, Member.section, Member.join_date,
        db.func.coalesce(db.func.sum(Payment.amount), 0).label("total_paid"),
        db.func.count(Payment.id).label("payment_count")
    ).outerjoin(Payment, Payment.member_id == Member.id).group_by(Member.id).order_by(Member.last_name).all()

    monthly_summary = db.session.query(
        db.func.strftime("%Y-%m", Payment.payment_date).label("month"),
        db.func.count(Payment.id).label("txns"),
        db.func.sum(Payment.amount).label("total")
    ).group_by("month").order_by(db.text("month DESC")).all()

    section_counts = db.session.query(Member.section, db.func.count(Member.id).label("count")).group_by(Member.section).all()
    attendance_summary = db.session.query(Attendance.status, db.func.count(Attendance.id).label("count")).group_by(Attendance.status).all()

    return render_template("reports.html", section="standard",
                           member_financials=member_financials, monthly_summary=monthly_summary,
                           section_counts=section_counts, attendance_summary=attendance_summary)


@bp.route("/export/members")
@login_required
def export_members():
    rows = Member.query.order_by(Member.last_name, Member.first_name).all()
    headers = ["id", "First Name", "Last Name", "Other Names", "Phone", "Email", "Section", "Join Date",
               "Address", "DOB", "State of Origin", "LGA", "NIN", "Qualification",
               "Country", "Passport No", "Zone", "Area", "Notes"]
    data = []
    for m in rows:
        data.append((m.id, m.first_name, m.last_name, m.other_names or "", m.phone or "", m.email or "",
                     m.section, m.join_date, m.address or "", m.dob or "", m.state_of_origin or "",
                     m.lga or "", m.nin_number or "", m.academic_qualification or "", m.country or "",
                     m.passport_number or "", m.zone or "", m.area or "", m.notes or ""))
    buf = export_excel("Members Export", headers, data, "members.xlsx", logo_path=_paths()[0])
    return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=members_{date.today().strftime('%Y%m%d')}.xlsx"})


@bp.route("/export/members/pdf")
@login_required
def export_members_pdf():
    rows = Member.query.order_by(Member.last_name, Member.first_name).all()
    headers = ["ID", "First Name", "Last Name", "Phone", "Email", "Section", "Join Date", "DOB", "State", "Country"]
    data = []
    for m in rows:
        data.append((m.id, m.first_name, m.last_name, m.phone or "", m.email or "",
                     m.section, m.join_date, m.dob or "", m.state_of_origin or "", m.country or "Nigeria"))
    buf = export_pdf("Members Report", headers, data, "members.pdf", logo_path=_paths())
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=members_{date.today().strftime('%Y%m%d')}.pdf"})


@bp.route("/export/payments")
@login_required
def export_payments():
    rows = db.session.execute(
        db.select(Payment.id, Member.first_name, Member.last_name, Payment.amount, Payment.payment_date, Payment.payment_for, Payment.notes)
        .join(Member).order_by(Payment.payment_date.desc())
    ).all()
    headers = ["id", "First Name", "Last Name", "Amount", "Date", "Payment For", "Notes"]
    data = [(r.id, r.first_name, r.last_name, float(r.amount), r.payment_date, r.payment_for or "", r.notes or "") for r in rows]
    buf = export_excel("Payments Export", headers, data, "payments.xlsx", logo_path=_paths()[0])
    return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=payments_{date.today().strftime('%Y%m%d')}.xlsx"})


@bp.route("/export/payments/pdf")
@login_required
def export_payments_pdf():
    rows = db.session.execute(
        db.select(Payment.id, Member.first_name, Member.last_name, Payment.amount, Payment.payment_date, Payment.payment_for, Payment.notes)
        .join(Member).order_by(Payment.payment_date.desc())
    ).all()
    headers = ["ID", "First Name", "Last Name", "Amount", "Date", "For", "Notes"]
    data = [(r.id, r.first_name, r.last_name, r.amount, r.payment_date, r.payment_for or "", r.notes or "") for r in rows]
    buf = export_pdf("Payments Report", headers, data, "payments.pdf", logo_path=_paths())
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=payments_{date.today().strftime('%Y%m%d')}.pdf"})


@bp.route("/export/attendance")
@login_required
def export_attendance():
    rows = db.session.execute(
        db.select(Attendance.id, Member.first_name, Member.last_name, Attendance.date, Attendance.status, Attendance.notes)
        .join(Member).order_by(Attendance.date.desc())
    ).all()
    headers = ["id", "First Name", "Last Name", "Date", "Status", "Notes"]
    data = [(r.id, r.first_name, r.last_name, r.date, r.status, r.notes or "") for r in rows]
    buf = export_excel("Attendance Export", headers, data, "attendance.xlsx", logo_path=_paths()[0])
    return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=attendance_{date.today().strftime('%Y%m%d')}.xlsx"})


@bp.route("/export/attendance/pdf")
@login_required
def export_attendance_pdf():
    rows = db.session.execute(
        db.select(Attendance.id, Member.first_name, Member.last_name, Attendance.date, Attendance.status, Attendance.notes)
        .join(Member).order_by(Attendance.date.desc())
    ).all()
    headers = ["ID", "First Name", "Last Name", "Date", "Status", "Notes"]
    data = [(r.id, r.first_name, r.last_name, r.date, r.status, r.notes or "") for r in rows]
    buf = export_pdf("Attendance Report", headers, data, "attendance.pdf", logo_path=_paths())
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=attendance_{date.today().strftime('%Y%m%d')}.pdf"})


@bp.route("/export/backup")
@login_required
def export_backup():
    wb = Workbook()
    thin = Border(
        left=Side(style="thin", color="D0D7E0"),
        right=Side(style="thin", color="D0D7E0"),
        top=Side(style="thin", color="D0D7E0"),
        bottom=Side(style="thin", color="D0D7E0")
    )
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")

    def write_sheet(ws, title, headers, rows):
        ws.title = title[:31]
        uc = [h.upper() for h in headers]
        for ci, h in enumerate(uc, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = thin
                cell.font = Font(size=9)
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(h) + 2, 14)

    members = Member.query.all()
    if members:
        cols = ["ID", "First Name", "Last Name", "Other Names", "Phone", "Email", "Section", "Join Date",
                "Address", "DOB", "State", "LGA", "NIN", "Qualification", "Country", "Passport No",
                "Photo", "Zone", "Area", "Notes"]
        data = []
        for m in members:
            data.append(tuple(getattr(m, c.lower().replace(" ", "_").replace(".",""), "") or "" for c in
                              ["ID", "First Name", "Last Name", "Other Names", "Phone", "Email", "Section",
                               "Join Date", "Address", "DOB", "State", "LGA", "NIN", "Qualification",
                               "Country", "Passport No", "Photo", "Zone", "Area", "Notes"]))
        write_sheet(wb.active, "Members", cols, data)

    payments = Payment.query.all()
    if payments:
        ws = wb.create_sheet()
        cols = ["ID", "Member ID", "Amount", "Date", "For", "Notes"]
        data = [(p.id, p.member_id, p.amount, p.payment_date, p.payment_for or "", p.notes or "") for p in payments]
        write_sheet(ws, "Payments", cols, data)

    attendance = Attendance.query.all()
    if attendance:
        ws = wb.create_sheet()
        cols = ["ID", "Member ID", "Date", "Status", "Notes"]
        data = [(a.id, a.member_id, a.date, a.status, a.notes or "") for a in attendance]
        write_sheet(ws, "Attendance", cols, data)

    songs = Song.query.all()
    if songs:
        ws = wb.create_sheet()
        cols = ["ID", "Title", "Composer", "Lyrics", "Audio File", "Upload Date", "Notes"]
        data = [(s.id, s.title, s.composer or "", s.lyrics or "", s.audio_file or "", s.upload_date, s.notes or "") for s in songs]
        write_sheet(ws, "Songs", cols, data)

    settings = Setting.query.all()
    if settings:
        ws = wb.create_sheet()
        cols = ["Key", "Value"]
        data = [(s.key, s.value or "") for s in settings]
        write_sheet(ws, "Settings", cols, data)

    users = User.query.all()
    if users:
        ws = wb.create_sheet()
        cols = ["ID", "Username", "Role", "Created At"]
        data = [(u.id, u.username, u.role, u.created_at) for u in users]
        write_sheet(ws, "Users", cols, data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=choir_full_backup_{date.today().strftime('%Y%m%d')}.xlsx"})


@bp.route("/reports/hierarchy/pdf")
@login_required
def reports_hierarchy_pdf():
    zone_f = request.args.get("zone", "")
    area_f = request.args.get("area", "")
    state_f = request.args.get("state", "")
    country_f = request.args.get("country", "")
    flat, hierarchy, grand = get_hierarchical_report(zone_f, area_f, state_f, country_f)
    choir_name = db.session.get(Setting, "choir_name").value if db.session.get(Setting, "choir_name") else "Choir"
    currency = db.session.get(Setting, "currency").value if db.session.get(Setting, "currency") else "$"
    buf = generate_hierarchy_pdf(flat, grand, choir_name, currency, logo_path=_paths())
    return Response(buf.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=hierarchy_report_{datetime.now().strftime('%Y%m%d')}.pdf"})


@bp.route("/reports/hierarchy/excel")
@login_required
def reports_hierarchy_excel():
    zone_f = request.args.get("zone", "")
    area_f = request.args.get("area", "")
    state_f = request.args.get("state", "")
    country_f = request.args.get("country", "")
    flat, hierarchy, grand = get_hierarchical_report(zone_f, area_f, state_f, country_f)
    choir_name = db.session.get(Setting, "choir_name").value if db.session.get(Setting, "choir_name") else "Choir"
    currency = db.session.get(Setting, "currency").value if db.session.get(Setting, "currency") else "$"
    buf = generate_hierarchy_excel(flat, grand, choir_name, currency, logo_path=_paths())
    return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=hierarchy_report_{datetime.now().strftime('%Y%m%d')}.xlsx"})


@bp.route("/reports/backups")
@login_required
def backups_list():
    backup_dir = os.path.join(current_app.root_path, "backups")
    os.makedirs(backup_dir, exist_ok=True)
    files = []
    for f in sorted(os.listdir(backup_dir), reverse=True):
        if f.startswith("auto_backup_") and f.endswith(".xlsx"):
            fpath = os.path.join(backup_dir, f)
            size = os.path.getsize(fpath)
            files.append({"name": f, "size": f"{size / 1024:.1f} KB", "ts": f.replace("auto_backup_", "").replace(".xlsx", "")})
    return render_template("backups.html", backups=files)


@bp.route("/reports/backups/<filename>")
@login_required
def backups_download(filename):
    backup_dir = os.path.join(current_app.root_path, "backups")
    return send_from_directory(backup_dir, filename, as_attachment=True)
