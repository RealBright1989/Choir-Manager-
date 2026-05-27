import os
import csv
import io
import re
import secrets
from datetime import datetime, date
from functools import wraps
from flask import session, flash, redirect, url_for, Response, request, current_app
from werkzeug.security import generate_password_hash
from io import BytesIO
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

from models import db, User, Setting, Member, Payment, Attendance, Song, PhoneVerification

try:
    from twilio.rest import Client as TwilioClient
    TWILIO_AVAILABLE = True
except Exception:
    TWILIO_AVAILABLE = False


def init_db_data():
    for key, val in [("choir_name", "8 Eton Natural Choir"), ("currency", "$"), ("dues_amount", "10"),
                     ("sms_provider", "log"), ("twilio_account_sid", ""), ("twilio_auth_token", ""), ("twilio_phone", ""),
                     ("smtp_host", ""), ("smtp_port", "587"), ("smtp_user", ""), ("smtp_password", ""), ("smtp_from", "")]:
        if not db.session.get(Setting, key):
            db.session.add(Setting(key=key, value=val))
    db.session.commit()
    if not User.query.first():
        db.session.add(User(username="admin", password_hash=generate_password_hash("admin123"),
                            role="admin", created_at=date.today().strftime("%Y-%m-%d")))
        db.session.commit()


def get_setting(key, default=""):
    s = db.session.get(Setting, key)
    return s.value if s else default


# ─── CSRF ──────────────────────────────────────────────────────────

def generate_csrf():
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_hex(32)
    return session["csrf_token"]


def validate_csrf():
    if os.environ.get("FLASK_ENV") == "testing":
        return True
    token = request.form.get("csrf_token")
    if not token or token != session.get("csrf_token"):
        flash("Invalid form token. Please try again.", "danger")
        return False
    return True


# ─── Auth decorator ────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in first.", "danger")
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)
    return decorated


# ─── Audit Log ──────────────────────────────────────────────────────

from models import AuditLog


def log_audit(action, entity_type, entity_id=None, details=""):
    try:
        entry = AuditLog(
            user_id=session.get("user_id", 0),
            username=session.get("username", "anonymous"),
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            details=str(details)[:500],
            ip_address=request.remote_addr or "",
            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )
        db.session.add(entry)
        db.session.commit()
    except Exception as e:
        logger.error(f"Audit log error: {e}")


# ─── Validation helpers ────────────────────────────────────────────

def validate_required(val, field_name):
    if not val or not val.strip():
        return f"{field_name} is required."
    return None


def validate_email(email):
    if email and not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        return "Invalid email format."
    return None


def validate_amount(amount):
    try:
        a = float(amount)
        if a <= 0:
            return "Amount must be greater than zero."
    except (ValueError, TypeError):
        return "Amount must be a valid number."
    return None


def validate_date(d, fmt="%Y-%m-%d"):
    try:
        datetime.strptime(d, fmt)
    except (ValueError, TypeError):
        return f"Invalid date format. Use {fmt}."
    return None


# ─── SMS / Phone Verification ─────────────────────────────────────

def get_sms_settings():
    keys = ["sms_provider", "twilio_account_sid", "twilio_auth_token", "twilio_phone"]
    return {k: get_setting(k) for k in keys}


def send_otp_sms(phone, otp):
    sms = get_sms_settings()
    provider = sms.get("sms_provider", "log")
    if provider == "twilio" and TWILIO_AVAILABLE:
        sid, token, from_phone = sms.get("twilio_account_sid", ""), sms.get("twilio_auth_token", ""), sms.get("twilio_phone", "")
        if sid and token and from_phone:
            try:
                TwilioClient(sid, token).messages.create(
                    body=f"Your {get_setting('choir_name', 'Choir')} verification code is: {otp}",
                    from_=from_phone, to=phone)
                logger.info(f"OTP sent via Twilio to {phone}")
                return True
            except Exception as e:
                logger.error(f"Twilio error: {e}")
                return False
    logger.info(f"[DEV MODE] OTP for {phone}: {otp}")
    return True


def generate_otp():
    return str(secrets.randbelow(900000) + 100000)


# ─── Email / Password Reset ─────────────────────────────────────────

def get_email_settings():
    keys = ["smtp_host", "smtp_port", "smtp_user", "smtp_password", "smtp_from"]
    return {k: get_setting(k) for k in keys}


def send_reset_email(email, username, reset_url):
    smtp = get_email_settings()
    host = smtp.get("smtp_host", "")
    if host and smtp.get("smtp_user") and smtp.get("smtp_password"):
        try:
            import smtplib
            from email.message import EmailMessage
            msg = EmailMessage()
            msg["Subject"] = f"Password Reset - {get_setting('choir_name', 'Choir')}"
            msg["From"] = smtp.get("smtp_from", host)
            msg["To"] = email
            msg.set_content(f"Hi {username},\n\nClick the link below to reset your password:\n\n{reset_url}\n\nThis link expires in 1 hour.\n\nIf you did not request this, ignore this email.")
            with smtplib.SMTP(host, int(smtp.get("smtp_port", 587))) as server:
                server.starttls()
                server.login(smtp["smtp_user"], smtp["smtp_password"])
                server.send_message(msg)
            logger.info(f"Password reset email sent to {email}")
            return True
        except Exception as e:
            logger.error(f"Email error: {e}")
            return False
    logger.info(f"[DEV MODE] Password reset link for {username}: {reset_url}")
    return True


def generate_reset_token():
    return secrets.token_urlsafe(48)


# ─── CSV Export Helper ─────────────────────────────────────────────

def export_csv(columns, rows, filename):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([getattr(row, c, "") if row is not None else "" for c in columns])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ─── Hierarchical Report Helper ────────────────────────────────────

def get_hierarchical_report(zone_filter="", area_filter="", state_filter="", country_filter=""):
    query = Member.query.outerjoin(Payment).add_columns(
        db.func.coalesce(db.func.sum(Payment.amount), 0).label("total_paid"),
        db.func.count(Payment.id).label("payment_count")
    ).group_by(Member.id).order_by(Member.country, Member.state_of_origin, Member.area, Member.zone, Member.last_name)

    if zone_filter:
        query = query.filter(Member.zone.like(f"%{zone_filter}%"))
    if area_filter:
        query = query.filter(Member.area.like(f"%{area_filter}%"))
    if state_filter:
        query = query.filter(Member.state_of_origin.like(f"%{state_filter}%"))
    if country_filter:
        query = query.filter(Member.country.like(f"%{country_filter}%"))

    results = query.all()
    hierarchy = {}
    for row in results:
        m = row[0] if hasattr(row, "__iter__") and not isinstance(row, Member) else row
        total_paid = getattr(row, "total_paid", 0) if not isinstance(row, Member) else 0
        payment_count = getattr(row, "payment_count", 0) if not isinstance(row, Member) else 0
        if not hasattr(m, "zone"):
            continue
        zone = m.zone or "Unassigned"
        area = m.area or "Unassigned"
        state = m.state_of_origin or "Unassigned"
        country = m.country or "Unassigned"
        hierarchy.setdefault(country, {}).setdefault(state, {}).setdefault(area, {}).setdefault(zone, [])
        hierarchy[country][state][area][zone].append((m, total_paid, payment_count))

    flat = []
    grand = {"members": 0, "payments": 0, "total": 0}
    for country in sorted(hierarchy):
        cdata = hierarchy[country]
        c_mem = sum(len(z) for s in cdata.values() for a in s.values() for z in a.values())
        c_pay = sum(pc for s in cdata.values() for a in s.values() for z in a.values() for _, _, pc in z)
        c_tot = sum(tp for s in cdata.values() for a in s.values() for z in a.values() for _, tp, _ in z)
        flat.append({"level": "country", "name": country, "members": c_mem, "payments": c_pay, "total": c_tot})
        grand["members"] += c_mem; grand["payments"] += c_pay; grand["total"] += c_tot
        for state in sorted(cdata):
            sdata = cdata[state]
            s_mem = sum(len(z) for a in sdata.values() for z in a.values())
            s_pay = sum(pc for a in sdata.values() for z in a.values() for _, _, pc in z)
            s_tot = sum(tp for a in sdata.values() for z in a.values() for _, tp, _ in z)
            flat.append({"level": "state", "name": state, "members": s_mem, "payments": s_pay, "total": s_tot})
            for area in sorted(sdata):
                adata = sdata[area]
                a_mem = sum(len(z) for z in adata.values())
                a_pay = sum(pc for z in adata.values() for _, _, pc in z)
                a_tot = sum(tp for z in adata.values() for _, tp, _ in z)
                flat.append({"level": "area", "name": area, "members": a_mem, "payments": a_pay, "total": a_tot})
                for zone in sorted(adata):
                    zdata = adata[zone]
                    z_mem = len(zdata)
                    z_pay = sum(pc for _, _, pc in zdata)
                    z_tot = sum(tp for _, tp, _ in zdata)
                    flat.append({"level": "zone", "name": zone, "members": z_mem, "payments": z_pay, "total": z_tot})
                    for m, tp, pc in zdata:
                        flat.append({"level": "member", "name": f"{m.first_name} {m.last_name}", "section": m.section, "phone": m.phone or "", "dob": m.dob or "", "members": 1, "payments": pc, "total": tp})
    return flat, hierarchy, grand


# ─── PDF Report Helper ─────────────────────────────────────────────

def generate_hierarchy_pdf(flat, grand, choir_name, currency, logo_path=None):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    if logo_path and os.path.exists(logo_path):
        pdf.image(logo_path, x=10, y=8, w=16)

    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(233, 69, 96)
    pdf.cell(0, 12, f"{choir_name}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(26, 26, 46)
    pdf.cell(0, 10, "Hierarchical Report (Zone > Area > State > International)", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_fill_color(26, 26, 46)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 10, f"  Grand Total: {grand['members']} Members | {grand['payments']} Payments | {currency}{grand['total']:.2f}", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    col_w = [10, 90, 30, 30, 30]
    headers = ["#", "Level / Name", "Members", "Payments", "Amount"]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(233, 69, 96)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 8, h, border=1, fill=True, align="C" if i > 1 else "L")
    pdf.ln()
    pdf.set_font("Helvetica", "", 8)
    level_colors = {"country": (15, 52, 96), "state": (26, 26, 46), "area": (22, 160, 133), "zone": (243, 156, 18), "member": (100, 100, 100)}
    level_prefix = {"country": "  ", "state": "    ", "area": "      ", "zone": "        ", "member": "          "}
    for idx, item in enumerate(flat, 1):
        pdf.set_text_color(*level_colors.get(item["level"], (0, 0, 0)))
        pdf.set_font("Helvetica", "B" if item["level"] != "member" else "", 8)
        name = level_prefix.get(item["level"], "") + item["name"]
        pdf.cell(col_w[0], 7, str(idx), border=1, align="C")
        pdf.cell(col_w[1], 7, name, border=1)
        pdf.cell(col_w[2], 7, str(item["members"]), border=1, align="C")
        pdf.cell(col_w[3], 7, str(item["payments"]), border=1, align="C")
        pdf.cell(col_w[4], 7, f"{currency}{item['total']:.2f}", border=1, align="R")
        pdf.ln()
    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf


def export_pdf(title, headers, rows, filename, logo_path=None):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    if logo_path and os.path.exists(logo_path):
        pdf.image(logo_path, x=10, y=8, w=16)

    pdf.ln(8)
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(233, 69, 96)
    pdf.cell(0, 14, title, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(130, 130, 130)
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)

    ncols = len(headers)
    page_w = pdf.w - 2 * pdf.l_margin
    col_w = [max(page_w / ncols, 18) for _ in range(ncols)]

    uc_headers = [h.upper() for h in headers]
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(26, 26, 46)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(uc_headers):
        align = "C" if i > 0 else "L"
        pdf.cell(col_w[i], 9, f" {h} ", border=1, fill=True, align=align)
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    for ri, row in enumerate(rows):
        if ri % 2 == 0:
            pdf.set_fill_color(244, 246, 249)
        else:
            pdf.set_fill_color(255, 255, 255)

        row_h = 8
        for ci, val in enumerate(row):
            s = str(val) if val is not None else ""
            if isinstance(val, float):
                s = f"{val:,.2f}"
            align = "C" if ci > 0 else "L"
            pdf.set_text_color(50, 50, 50)
            pdf.cell(col_w[ci], row_h, f" {s} ", border=1, fill=True, align=align)
        pdf.ln()

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf


# ─── Excel Report Helper ───────────────────────────────────────────

def generate_hierarchy_excel(flat, grand, choir_name, currency, logo_path=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hierarchy Report"

    if logo_path and os.path.exists(logo_path):
        img = XlImage(logo_path)
        img.width = 60
        img.height = 60
        ws.add_image(img, "A1")

    title_font = Font(bold=True, size=16, color="E94560")
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    alt_fill = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
    thin = Border(
        left=Side(style="thin", color="D0D7E0"),
        right=Side(style="thin", color="D0D7E0"),
        top=Side(style="thin", color="D0D7E0"),
        bottom=Side(style="thin", color="D0D7E0")
    )
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"{choir_name} - Hierarchical Report"
    c.font = title_font
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"Zone > Area > State > International  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font = Font(size=9, color="666666")
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:H4")
    c = ws["A4"]
    c.value = f"Grand Total: {grand['members']} Members  |  {grand['payments']} Payments  |  {currency}{grand['total']:.2f}"
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    c.alignment = Alignment(horizontal="left")

    headers = ["#", "LEVEL / NAME", "SECTION", "PHONE", "DATE OF BIRTH", "MEMBERS", "PAYMENTS", "AMOUNT"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    level_prefix = {"country": "  ", "state": "    ", "area": "      ", "zone": "        ", "member": "          "}
    for ri, item in enumerate(flat, 7):
        name = level_prefix.get(item["level"], "") + item["name"]
        row_num = ri - 6
        is_member = item["level"] == "member"
        phone_val = item.get("phone", "") if is_member else ""
        section_val = item.get("section", "") if is_member else ""
        dob_val = item.get("dob", "") if is_member else ""
        members_val = item["members"]
        payments_val = item["payments"]
        amount_val = item["total"]

        vals = [row_num, name, section_val, phone_val, dob_val, members_val, payments_val, amount_val]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin
            cell.font = Font(
                bold=not is_member, size=9,
                color={"country": "0F3460", "state": "1A1A2E", "area": "16A085", "zone": "F39C12", "member": "555555"}.get(item["level"], "000000")
            )
            if ci >= 6:
                cell.alignment = Alignment(horizontal="center")
            if ci == 8:
                cell.number_format = '#,##0.00'
        if (ri - 7) % 2:
            for ci in range(1, 9):
                ws.cell(row=ri, column=ci).fill = alt_fill

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 18

    for row in ws.iter_rows(min_row=6, max_row=len(flat)+6, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_excel(title, headers, rows, filename, logo_path=None):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    if logo_path and os.path.exists(logo_path):
        img = XlImage(logo_path)
        img.width = 60
        img.height = 60
        ws.add_image(img, "A1")

    title_font = Font(bold=True, size=14, color="E94560")
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    alt_fill = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
    thin = Border(
        left=Side(style="thin", color="D0D7E0"),
        right=Side(style="thin", color="D0D7E0"),
        top=Side(style="thin", color="D0D7E0"),
        bottom=Side(style="thin", color="D0D7E0")
    )

    last_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title
    c.font = title_font
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    uc_headers = [h.upper() for h in headers]
    for ci, h in enumerate(uc_headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    for ri, row in enumerate(rows, 4):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin
            cell.font = Font(size=9, color="333333")
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="center")
            elif isinstance(val, int):
                cell.alignment = Alignment(horizontal="center")
        if (ri - 4) % 2:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).fill = alt_fill

    col_widths = {0: 6}
    for ci, h in enumerate(headers):
        col_widths[ci] = max(len(h) + 2, 12)
        if h.lower() in ("amount", "total", "fee"):
            col_widths[ci] = 16
        elif h.lower() in ("phone", "email", "address", "notes"):
            col_widths[ci] = 22
        elif h.lower() in ("name", "first_name", "last_name", "title", "composer"):
            col_widths[ci] = 18
    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci + 1)].width = w

    for row in ws.iter_rows(min_row=3, max_row=len(rows) + 3, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
